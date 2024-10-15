#######################
# @author: RaffaDNDM
# @date:   2022-09-09
#######################

import xml.etree.ElementTree as ET
import shutil
from bs4 import BeautifulSoup
import base64
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from alive_progress import alive_bar
import argparse
import os
from datetime import datetime

#Security Headers to be searched in response headers
SECURITY_HEADERS = {'x-frame-options': set(), 
                    'x-content-type-options': set(),
                    'strict-transport-security': set(),
                    'content-security-policy': set(),
                    'referrer-policy': set(),
                    'x-xss-protection': set(),
                    'expect-ct': set(),
                    'permissions-policy': set(),
                    'cross-origin-embedder-policy': set(),
                    'cross-origin-resource-policy': set(),
                    'cross-origin-opener-policy': set()                   
                   }

FINGERPRINTING_HEADERS = {'via': [],
                          'server': [],
                          'x-powered-by': [],
                          'x-aspnet-version': []
                         }

#Header that set options for cookie (to be searched in response headers)
COOKIE_HEADER='set-cookie'

COOKIE_ATTRIBUTES = { 'expires': '',
                      'max-age': '',
                      'domain': '',
                      'path': '',
                      'samesite': ''
                    }

COOKIE_ATTRIBUTES_BOOL = { 'secure': False,
                           'httponly': False
                         }

#Hosts security headers in responses
HOSTS_SEC_HEADERS = {}
#Cookie options for each host
COOKIES_OPTIONS = {}
#Fingerprinting info
FINGERPRINT = {}

def bs4_parser(input_file, output_folder, customer_name, current_time):
    """
    Parse Burp HTTP History XML file.

    Parameters
    ----------
    input_file (str): Name of HTTP History XML file (e.g. HTTPHistory.xml)

    output_folder (str): Path of the output folder, used to create results files.

    customer_name (str): Name of the customer, owner of the domains analysed.

    current_time (str): Time when the analysis is performed (it will be used to
                        define results filenames)
    """
    
    #Read input XML file
    print("\nReading input file...", end='\t')
    infile = open(input_file,"r")
    contents = infile.read()
    soup = BeautifulSoup(contents,'xml')    
    print("DONE", end='\n\n')

    #Find all items (request/response)
    items = soup.find_all('item')
    NUM_ITEMS = len(items)
    
    #Analysis of all the items
    print("Analysing items")
    with alive_bar(NUM_ITEMS) as bar:
        
        #For each item extract host, method, path, response
        for x in items:
            response_item = x.find('response')

            response = ''
            if response_item:
                response = response_item.get_text()

            if response!="":
                host=x.find('host').get_text()
                method = x.find('method').get_text()
                path = x.find('path').get_text()
                
                #Take & decode base64 request
                request_item = x.find('request')
                request = request_item.get_text()
                request_base64_bytes = request.encode('ISO-8859-1')
                request_bytes = base64.b64decode(request_base64_bytes)

                #Decode base64 response
                response_base64_bytes = response.encode('ISO-8859-1')
                response_bytes = base64.b64decode(response_base64_bytes)
                
                #ANalysis of response headers
                parse_HTTP(host, method, path, response_bytes.decode('ISO-8859-1'), request_bytes.decode('ISO-8859-1'))

            #Update progress bar status
            bar()

    #Create CSV files with results
    print('\nCompiling CSV files...', end='\t')
    generate_csv(output_folder, customer_name, current_time)
    print("DONE", end='\n\n')
    
    #Create XLSX file from template and compile it
    print('Compiling XLSX file...', end='\t')
    compile_xlsx(output_folder, customer_name, current_time)
    print("DONE", end='\n\n')

def generate_csv(output_folder, customer_name, current_time):
    """
    Create CSV results files.

    Parameters
    ----------
    output_folder (str): Path of the output folder, used to create results files.

    customer_name (str): Name of the customer, owner of the domains analysed.

    current_time (str): Time when the analysis is performed (it will be used to
                        define results filenames)
    """

    global HOSTS_SEC_HEADERS
    global COOKIES_OPTIONS
    global FINGERPRINT

    #Define output CSV filenames
    sh_filename = os.path.join(output_folder, current_time + '-' + customer_name +'-Security_Headers.csv')
    msh_filename = os.path.join(output_folder, current_time + '-' + customer_name +'-Missing_Security_Headers.csv')
    finger_filename = os.path.join(output_folder, current_time + '-' + customer_name +'-Fingerprint_HTTP_Headers.csv')
    cookie_filename = os.path.join(output_folder, current_time + '-' + customer_name +'-Cookies.csv')

    #Open output CSV files for analysis of security headers
    with open(sh_filename, 'w') as sh, open(msh_filename, 'w') as msh:
        sh.write(f'Host,Method,Path,Header,Value\n')
        msh.write(f'Host,Method,Path,Missing Header\n')

        #Collect information stored after parsing HTTP response headers
        for host in HOSTS_SEC_HEADERS:
            for method in HOSTS_SEC_HEADERS[host]:
                for path in HOSTS_SEC_HEADERS[host][method]:
                    for h in HOSTS_SEC_HEADERS[host][method][path]:
                        if len(HOSTS_SEC_HEADERS[host][method][path][h]):
                            for value in HOSTS_SEC_HEADERS[host][method][path][h]:
                                
                                row_element={ 'Host': host,
                                            'Method': method,
                                            'Path': path,
                                            'Header': h,
                                            'Value': value
                                            } 

                                #Write Security headers values to output file for each (host, method, path) tuple
                                sh.write(f'{host},{method},{path},{h},{value}\n')
                        else:
                            row_element={ 'Host': host,
                                        'Method': method,
                                        'Path': path,
                                        'Missing Header': h,
                                        }
                            
                            #Write Security header names that are missing for each (host, method, path) tuple
                            msh.write(f'{host},{method},{path},{h}\n')

    #Open output CSV files for analysis of Set-Cookie options
    with open(finger_filename, 'w', encoding="utf-8") as fd:
        fd.write(f'Host,Header Name,Header Value\n')
        for host in FINGERPRINT:
            for h in FINGERPRINT[host]:
                if len(FINGERPRINT[host][h]):
                    for item in FINGERPRINT[host][h]:
                        fd.write(f'{host},{h},{item[0]}\n')

    #Open output CSV files for analysis of Set-Cookie options
    with open(cookie_filename, 'w') as fd:
        fd.write(f'Host,Path,Set-Cookie options\n')
        
        #Write Set-Cookie options each (host, path) couple
        for host in COOKIES_OPTIONS:
            for path in COOKIES_OPTIONS[host]:
                fd.write(f'{host},{path},{COOKIES_OPTIONS[host][path]}\n')


def compile_xlsx(output_folder, customer_name, current_time):
    """
    Create XLSX results file.

    Parameters
    ----------
    output_folder (str): Path of the output folder, used to create results files.

    customer_name (str): Name of the customer, owner of the domains analysed.

    current_time (str): Time when the analysis is performed (it will be used to
                        define results filenames)
    """

    global HOSTS_SEC_HEADERS
    global COOKIES_OPTIONS
    global FINGERPRINT

    #Define output XLSX filename
    results_filename = os.path.join(output_folder, current_time + '-' + customer_name +'-Security_Analysis.xlsx')
    
    #Create XLSX file from Template file
    shutil.copy("Template.xlsx", results_filename)
    
    #Open XLSX results file
    wb = load_workbook(results_filename)

    #Coordinates in XLSX file w.r.t. sheet name
    coordinates = { 'Security Headers': { 'HEADER ROW': wb['Security Headers'].max_row,
                        'xlsx row'  : wb['Security Headers'].max_row+1,
                        'MIN COL' : wb['Security Headers'].min_column,
                        'MAX COL' : wb['Security Headers'].max_column            
                    },
                    'Missing Security Headers': { 'HEADER ROW': wb['Missing Security Headers'].max_row,
                        'xlsx row'  :  wb['Missing Security Headers'].max_row+1,
                        'MIN COL' :  wb['Missing Security Headers'].min_column,
                        'MAX COL' :  wb['Missing Security Headers'].max_column        
                    },
                    'HTTP Headers Fingerprint': { 'HEADER ROW': wb['HTTP Headers Fingerprint'].max_row,
                        'xlsx row'  :  wb['HTTP Headers Fingerprint'].max_row+1,
                        'MIN COL' :  wb['HTTP Headers Fingerprint'].min_column,
                        'MAX COL' :  wb['HTTP Headers Fingerprint'].max_column        
                    },
                    'Cookies': { 'HEADER ROW': wb['Cookies'].max_row,
                        'xlsx row'  :  wb['Cookies'].max_row+1,
                        'MIN COL' :  wb['Cookies'].min_column,
                        'MAX COL' :  wb['Cookies'].max_column        
                    }
                  }
    
    #Collect information stored after parsing HTTP response headers
    for host in HOSTS_SEC_HEADERS:
        for method in HOSTS_SEC_HEADERS[host]:
            for path in HOSTS_SEC_HEADERS[host][method]:
                for h in HOSTS_SEC_HEADERS[host][method][path]:
                    if len(HOSTS_SEC_HEADERS[host][method][path][h]):
                        for value in HOSTS_SEC_HEADERS[host][method][path][h]:
                            
                            row_element={ 'Host': host,
                                        'Method': method,
                                        'Path': path,
                                        'Header': h,
                                        'Value': value
                                        }

                            #Write Security headers values to output file for each (host, method, path) tuple
                            for i in range(coordinates['Security Headers']['MIN COL'],coordinates['Security Headers']['MAX COL']+1):
                                cell=wb['Security Headers'][get_column_letter(i)+str(coordinates['Security Headers']['xlsx row'])]
                                cell.value=row_element[wb['Security Headers'][get_column_letter(i)+str(coordinates['Security Headers']['HEADER ROW'])].value]

                            coordinates['Security Headers']['xlsx row']+=1
                    else:
                        
                        row_element={ 'Host': host,
                                      'Method': method,
                                      'Path': path,
                                      'Missing Header': h,
                                    }

                        #Write Security header names that are missing for each (host, method, path) tuple
                        for i in range(coordinates['Missing Security Headers']['MIN COL'],coordinates['Missing Security Headers']['MAX COL']+1):
                            cell=wb['Missing Security Headers'][get_column_letter(i)+str(coordinates['Missing Security Headers']['xlsx row'])]
                            cell.value=row_element[wb['Missing Security Headers'][get_column_letter(i)+str(coordinates['Missing Security Headers']['HEADER ROW'])].value]

                        coordinates['Missing Security Headers']['xlsx row']+=1

    
    for host in FINGERPRINT:
        for h in FINGERPRINT[host]:
            if len(FINGERPRINT[host][h]):
                for item in FINGERPRINT[host][h]:
                    row_element = { 'Host': host,
                                    'Header Name': h,
                                    'Header Value': item[0],
                                    'Request': item[1],
                                    'Response': item[2]
                                }
                    
                    for i in range(coordinates['HTTP Headers Fingerprint']['MIN COL'],coordinates['HTTP Headers Fingerprint']['MAX COL']+1):
                            cell=wb['HTTP Headers Fingerprint'][get_column_letter(i)+str(coordinates['HTTP Headers Fingerprint']['xlsx row'])]
                            cell.value=row_element[wb['HTTP Headers Fingerprint'][get_column_letter(i)+str(coordinates['HTTP Headers Fingerprint']['HEADER ROW'])].value]

                    coordinates['HTTP Headers Fingerprint']['xlsx row']+=1

    #Open output CSV files for analysis of Set-Cookie options
    for host in COOKIES_OPTIONS:
        for path in COOKIES_OPTIONS[host]:

            row_element = { 'Host': host,
                            'Dir': path,
                            'Cookie Name': COOKIES_OPTIONS[host][path]['cookie name'],
                            'Cookie Value': COOKIES_OPTIONS[host][path]['cookie value'],
                            'Expires': COOKIES_OPTIONS[host][path]['expires'],
                            'Max-Age': COOKIES_OPTIONS[host][path]['max-age'],
                            'Domain': COOKIES_OPTIONS[host][path]['domain'],
                            'Path': COOKIES_OPTIONS[host][path]['path'],
                            'SameSite': COOKIES_OPTIONS[host][path]['samesite'],
                            'Secure': COOKIES_OPTIONS[host][path]['secure'],
                            'HTTPOnly': COOKIES_OPTIONS[host][path]['httponly'] }

            #Write Set-Cookie options each (host, path) couple
            for i in range(coordinates['Cookies']['MIN COL'],coordinates['Cookies']['MAX COL']+1):
                cell=wb['Cookies'][get_column_letter(i)+str(coordinates['Cookies']['xlsx row'])]
                cell.value=row_element[wb['Cookies'][get_column_letter(i)+str(coordinates['Cookies']['HEADER ROW'])].value]

            coordinates['Cookies']['xlsx row']+=1

    #Save changes to XLSX file
    wb.save(results_filename)
    
def parse_cookies(cookie_value: str):
    global COOKIE_ATTRIBUTES
    global COOKIE_ATTRIBUTES_BOOL
    
    attributes = cookie_value.replace(' ', '').split(';')

    cookie_name, cookie_value = attributes[0].strip().split('=', 1)
    
    cookie_spec = { 'cookie name': cookie_name,
                    'cookie value': cookie_value,
                  }
                  
    cookie_spec.update(COOKIE_ATTRIBUTES.copy())
    cookie_spec.update(COOKIE_ATTRIBUTES_BOOL.copy())

    for att in attributes[1:]:
        
        if '=' in att:
            att_name, att_value = att.split('=', 1)

            print(att_name)

            if att_name.lower() in COOKIE_ATTRIBUTES:
                cookie_spec[att_name.lower()] = att_value
    
        elif att.lower() in COOKIE_ATTRIBUTES_BOOL:
            cookie_spec[att.lower()] = True
            
    return cookie_spec

def parse_HTTP(host: str, method: str, path: str, response: str, request: str):
    """
    Parse an HTTP response.

    Parameters
    ----------
    host (str): Name of the host that replies to the current response.

    method (str): Method used to contact the host that replies to the current response.
    
    path (str): Path on which the request was performed.

    response (str): Response to be parsed.

    request (str): Request that generated the response to be parsed.
    """
    
    global HOSTS_SEC_HEADERS
    global SECURITY_HEADERS
    global COOKIES_OPTIONS
    global COOKIE_HEADER
    global FINGERPRINTING_HEADERS
    global FINGERPRINT

    #Add host parsed to list of hosts
    if host not in HOSTS_SEC_HEADERS:
        HOSTS_SEC_HEADERS[host]=dict()
        COOKIES_OPTIONS[host]=dict()
        FINGERPRINT[host] = FINGERPRINTING_HEADERS.copy()

    #Add method parsed to list of methods used for an host
    if method not in HOSTS_SEC_HEADERS[host].keys():
        HOSTS_SEC_HEADERS[host][method]=dict()
        
    #Add path parsed to (host, method) couple
    if path not in HOSTS_SEC_HEADERS[host][method].keys():
        HOSTS_SEC_HEADERS[host][method][path]=SECURITY_HEADERS.copy()

    #Split headers from body and store only headers
    headers = response.split('\r\n\r\n', 1)[0]
    
    #Define a dictionary of all (Header Name: Header value) couples of the response
    headers_dict = dict()
    for x in headers.split("\r\n")[1:]:
        split_x = x.split(": ", 1)
        #Store headers value as case insensitive and without spaces
        headers_dict[split_x[0].lower().replace(' ', '')] = split_x[1]
    
    #For each header, add its value to the list of values related to each security header
    #for all the (host, method, path) tuples
    for h in headers_dict:
        if h in SECURITY_HEADERS:
            HOSTS_SEC_HEADERS[host][method][path][h].add(headers_dict[h])

        if h in FINGERPRINTING_HEADERS:
            FINGERPRINT[host][h].append((headers_dict[h], request, headers+'\r\n\r\n...'))
        
        if h==COOKIE_HEADER:
            COOKIES_OPTIONS[host][path] = dict()
            COOKIES_OPTIONS[host][path] = parse_cookies(headers_dict[h])

def input_parameters():
    """
    Parse command line parameters.

    Return
    ----------
    input_file (str): Name of HTTP History XML file (e.g. HTTPHistory.xml)

    output_folder (str): Path of the output folder, used to create results files.

    customer_name (str): Name of the customer, owner of the domains analysed.
    """    
    
    #Define argument parser
    parser = argparse.ArgumentParser()

    #Create command line arguments
    parser.add_argument('--input', '-in', dest='input_file', help='Input Burp HTTPHistory XML file.')
    parser.add_argument('--output', '-out', dest='output_folder', help='Existing/To Be Created output folder')
    parser.add_argument('-client', '-customer', dest='customer_name', help='Client owner of the domains to be tested')

    #Parse command line arguments
    args = parser.parse_args()

    #Ask user input filename if not provided as command line argument
    input_file=args.input_file
    while (not input_file) or input_file=='' or (not os.path.exists(input_file)):
        print('_Input file_')
        input_file = input()
    
    #Change output folder if the user provided it on command line (By default: '.')
    output_folder ='.'
    if args.output_folder:
        output_folder = args.output_folder

        if not os.path.exists(output_folder):
            os.mkdir(output_folder)

    #Ask Customer name, owner of the domains analysed, if not provided as command line argument
    customer_name = args.customer_name
    while (not customer_name) or customer_name=='':
        print('_Customer name_')
        customer_name = input()

    return input_file, output_folder, customer_name

def main():
    #Input parameters file
    input_file, output_folder, customer_name = input_parameters()

    #Current time
    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d_%H-%M-%S")

    #Parsing of Burp HTTP history
    bs4_parser(input_file, output_folder, customer_name, current_time)

if __name__=="__main__":
    main()